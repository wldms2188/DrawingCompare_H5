"""DrawingCompare H5 desktop GUI entry point."""
from __future__ import annotations

import threading
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import cv2
import numpy as np

from version import APP_NAME, VERSION
from core.image_loader import ImageLoader
from core.auto_align import AutoAlign
from core.change_detector import ChangeDetector


class DrawingCompareApp(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title(f"{APP_NAME} {VERSION}")
        self.geometry("900x650")
        self.minsize(760, 560)
        self.before_dir = tk.StringVar()
        self.after_dir = tk.StringVar()
        self.output_dir = tk.StringVar(value=str(Path(__file__).resolve().parent / "output"))
        self.status = tk.StringVar(value="Before / After 폴더를 선택하세요.")
        self._build()

    def _build(self):
        root = ttk.Frame(self, padding=18)
        root.pack(fill="both", expand=True)
        ttk.Label(root, text=f"{APP_NAME} {VERSION}", font=("Segoe UI", 20, "bold")).pack(anchor="w")
        ttk.Label(root, text="Engineering Drawing PDF Compare", font=("Segoe UI", 10)).pack(anchor="w", pady=(0, 18))

        box = ttk.LabelFrame(root, text="입력", padding=12)
        box.pack(fill="x")
        self._path_row(box, "Before", self.before_dir, False)
        self._path_row(box, "After", self.after_dir, False)
        self._path_row(box, "결과 폴더", self.output_dir, True)

        self.run_btn = ttk.Button(root, text="도면 비교 시작", command=self.start)
        self.run_btn.pack(anchor="e", pady=14)

        self.progress = ttk.Progressbar(root, mode="indeterminate")
        self.progress.pack(fill="x", pady=(0, 10))
        ttk.Label(root, textvariable=self.status).pack(anchor="w")

        logbox = ttk.LabelFrame(root, text="처리 로그", padding=8)
        logbox.pack(fill="both", expand=True, pady=(12, 0))
        self.log = tk.Text(logbox, wrap="word", state="disabled", font=("Consolas", 9))
        self.log.pack(side="left", fill="both", expand=True)
        scroll = ttk.Scrollbar(logbox, command=self.log.yview)
        scroll.pack(side="right", fill="y")
        self.log.configure(yscrollcommand=scroll.set)

    def _path_row(self, parent, label, variable, output):
        row = ttk.Frame(parent)
        row.pack(fill="x", pady=4)
        ttk.Label(row, text=label, width=10).pack(side="left")
        ttk.Entry(row, textvariable=variable).pack(side="left", fill="x", expand=True, padx=6)
        ttk.Button(row, text="찾아보기", command=lambda: self.choose(variable, output)).pack(side="right")

    def choose(self, variable, output=False):
        path = filedialog.askdirectory()
        if path:
            variable.set(path)

    def write_log(self, text):
        self.after(0, self._write_log, text)

    def _write_log(self, text):
        self.log.configure(state="normal")
        self.log.insert("end", text + "\n")
        self.log.see("end")
        self.log.configure(state="disabled")

    def start(self):
        before, after = Path(self.before_dir.get()), Path(self.after_dir.get())
        output = Path(self.output_dir.get())
        if not before.is_dir() or not after.is_dir():
            messagebox.showwarning("입력 확인", "Before와 After 폴더를 모두 선택하세요.")
            return
        self.run_btn.configure(state="disabled")
        self.progress.start(12)
        self.status.set("PDF를 읽고 있습니다...")
        threading.Thread(target=self._worker, args=(before, after, output), daemon=True).start()

    def _worker(self, before_dir, after_dir, output_dir):
        try:
            output_dir.mkdir(parents=True, exist_ok=True)
            loader = ImageLoader()
            before_docs = loader.load_folder(before_dir)
            after_docs = loader.load_folder(after_dir)
            if not before_docs or not after_docs:
                raise RuntimeError("Before 또는 After 폴더에 읽을 수 있는 PDF가 없습니다.")
            self.write_log(f"Before PDF {len(before_docs)}개 / After PDF {len(after_docs)}개")

            pairs = self._match_documents(before_docs, after_docs)
            aligner = AutoAlign()
            detector = ChangeDetector()
            rows = []
            total = sum(min(len(a.pages), len(b.pages)) for a, b in pairs)
            done = 0
            capture_dir = output_dir / "captures"
            capture_dir.mkdir(parents=True, exist_ok=True)

            for pair_no, (bd, ad) in enumerate(pairs, 1):
                page_pairs = self._match_pages(bd.pages, ad.pages)
                self.write_log(f"문서 {pair_no}: {bd.filename} ↔ {ad.filename}, {len(page_pairs)}페이지")
                for bp, ap in page_pairs:
                    done += 1
                    self.after(0, lambda d=done, t=max(total, 1): self.status.set(f"비교 중... {d}/{t}"))
                    before_img = bp.image
                    after_img = ap.image
                    try:
                        aligned = aligner.align(before_img, after_img)
                    except Exception as exc:
                        self.write_log(f"  정렬 경고 p{bp.page_index + 1}: {exc}")
                        aligned = after_img
                    try:
                        result = detector.detect(before_img, aligned)
                    except Exception as exc:
                        self.write_log(f"  검출 오류 p{bp.page_index + 1}: {exc}")
                        continue
                    for region_no, region in enumerate(result.regions, 1):
                        stem = f"D{pair_no:02d}_P{bp.page_index + 1:03d}_R{region_no:03d}"
                        old_path = capture_dir / f"{stem}_before.png"
                        new_path = capture_dir / f"{stem}_after.png"
                        cv2.imwrite(str(old_path), self._to_bgr(region.old_crop))
                        cv2.imwrite(str(new_path), self._to_bgr(region.new_crop))
                        rows.append({
                            "No": len(rows) + 1, "Before PDF": bd.filename, "After PDF": ad.filename,
                            "Before Page": bp.page_index + 1, "After Page": ap.page_index + 1,
                            "Type": region.region_type, "Confidence": round(region.confidence, 3),
                            "X": region.x, "Y": region.y, "Width": region.width, "Height": region.height,
                            "Change Ratio": round(region.change_ratio, 4),
                            "Before Image": str(old_path), "After Image": str(new_path),
                        })
            report = output_dir / "DrawingCompare_H5_Result.xlsx"
            self._write_excel(rows, report)
            self.write_log(f"완료: 변경영역 {len(rows)}개")
            self.after(0, lambda: self._finished(report, len(rows)))
        except Exception as exc:
            self.write_log(f"ERROR: {exc}")
            self.after(0, lambda e=str(exc): self._failed(e))

    @staticmethod
    def _to_bgr(img):
        if img is None:
            return np.full((80, 150, 3), 255, dtype=np.uint8)
        arr = np.asarray(img)
        if arr.ndim == 2:
            return cv2.cvtColor(arr, cv2.COLOR_GRAY2BGR)
        if arr.shape[2] == 4:
            return cv2.cvtColor(arr, cv2.COLOR_RGBA2BGR)
        return arr

    @staticmethod
    def _match_documents(before_docs, after_docs):
        remaining = list(after_docs)
        pairs = []
        for bd in before_docs:
            if not remaining:
                break
            best = min(remaining, key=lambda ad: DrawingCompareApp._doc_score(bd, ad))
            pairs.append((bd, best))
            remaining.remove(best)
        return pairs

    @staticmethod
    def _doc_score(a, b):
        name_a = Path(a.filename).stem.lower()
        name_b = Path(b.filename).stem.lower()
        name_penalty = 0 if name_a == name_b else 1
        page_penalty = abs(a.page_count - b.page_count)
        return name_penalty * 10 + page_penalty

    @staticmethod
    def _match_pages(before_pages, after_pages):
        remaining = list(after_pages)
        result = []
        for bp in before_pages:
            if not remaining:
                break
            def score(ap):
                ar1, ar2 = bp.aspect_ratio, ap.aspect_ratio
                return abs(ar1 - ar2) + abs(bp.width * bp.height - ap.width * ap.height) / max(bp.width * bp.height, 1)
            ap = min(remaining, key=score)
            result.append((bp, ap))
            remaining.remove(ap)
        return result

    @staticmethod
    def _write_excel(rows, path):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.drawing.image import Image as XLImage
        wb = Workbook()
        ws = wb.active
        ws.title = "Drawing Compare"
        headers = ["No", "Before PDF", "After PDF", "Before Page", "After Page", "Type", "Confidence", "X", "Y", "Width", "Height", "Change Ratio", "Before Image", "After Image"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(1, col, h); c.font = Font(bold=True); c.alignment = Alignment(horizontal="center")
        for r, item in enumerate(rows, 2):
            for c, h in enumerate(headers, 1): ws.cell(r, c, item.get(h, ""))
            ws.row_dimensions[r].height = 90
            for col, key in ((13, "Before Image"), (14, "After Image")):
                p = item.get(key)
                if p and Path(p).exists():
                    img = XLImage(p); img.width = 150; img.height = 80; ws.add_image(img, f"{chr(64 + col)}{r}")
        for col in range(1, len(headers) + 1): ws.column_dimensions[chr(64 + col)].width = 18
        wb.save(path)

    def _finished(self, report, count):
        self.progress.stop(); self.run_btn.configure(state="normal")
        self.status.set(f"완료 — 변경영역 {count}개 / Excel 저장 완료")
        messagebox.showinfo("비교 완료", f"비교가 완료되었습니다.\n\n변경영역: {count}개\n결과: {report}")

    def _failed(self, error):
        self.progress.stop(); self.run_btn.configure(state="normal")
        self.status.set("오류가 발생했습니다.")
        messagebox.showerror("비교 오류", error)


if __name__ == "__main__":
    DrawingCompareApp().mainloop()
