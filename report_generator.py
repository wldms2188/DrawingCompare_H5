"""
DrawingCompare H5
report_generator.py
 
역할
------------------------------------------------------------
도면 비교 결과를 Excel 보고서로 생성한다.
 
보고서에는 다음 정보를 포함한다.
 
1. 파일명
2. Before / After 페이지
3. 변경 영역 ID
4. 변경 유형
5. Before 내용
6. After 내용
7. 신뢰도
8. 변경 사유
9. Before 이미지
10. After 이미지
11. 변경 위치
"""
 
from __future__ import annotations
 
from pathlib import Path
from typing import List, Optional
 
from datetime import datetime
 
import os
 
from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    Alignment,
    Border,
    Side,
)
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
 
from core.change_detector import (
    ChangeDetectionResult,
    ChangeRegion,
)
 
from core.ocr_detector import (
    TextChange,
)
 
 
# ============================================================
# REPORT GENERATOR
# ============================================================
 
class ReportGenerator:
 
    def __init__(
        self,
        output_dir="output"
    ):
 
        self.output_dir = Path(
            output_dir
        )
 
        self.output_dir.mkdir(
            parents=True,
            exist_ok=True
        )
 
        # ----------------------------------------------------
        # Excel 기본 설정
        # ----------------------------------------------------
 
        self.title_font = Font(
            bold=True,
            size=14
        )
 
        self.header_font = Font(
            bold=True
        )
 
        self.normal_alignment = Alignment(
            vertical="center",
            wrap_text=True
        )
 
        self.center_alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
 
        self.border = Border(
            left=Side(
                style="thin"
            ),
            right=Side(
                style="thin"
            ),
            top=Side(
                style="thin"
            ),
            bottom=Side(
                style="thin"
            ),
        )
 
 
    # ========================================================
    # CREATE WORKBOOK
    # ========================================================
 
    def create_workbook(
        self
    ) -> Workbook:
        """
        기본 Excel Workbook을 생성한다.
        """
 
        workbook = Workbook()
 
        # 기본 Sheet
        summary_sheet = (
            workbook.active
        )
 
        summary_sheet.title = (
            "Summary"
        )
 
        # 변경점 Sheet
        workbook.create_sheet(
            "Changes"
        )
 
        # 정렬 Sheet
        workbook.create_sheet(
            "Alignment"
        )
 
        return workbook
 
 
    # ========================================================
    # SUMMARY SHEET
    # ========================================================
 
    def create_summary_sheet(
        self,
        workbook: Workbook,
        before_file_count: int,
        after_file_count: int,
        matched_file_count: int,
        total_page_count: int,
        total_change_count: int,
        output_path: Optional[str] = None,
    ):
        """
        Summary Sheet를 생성한다.
        """
 
        sheet = workbook[
            "Summary"
        ]
 
        # ----------------------------------------------------
        # 제목
        # ----------------------------------------------------
 
        sheet["A1"] = (
            "DrawingCompare H5"
        )
 
        sheet["A1"].font = (
            self.title_font
        )
 
        sheet.merge_cells(
            "A1:D1"
        )
 
        # ----------------------------------------------------
        # 생성 시간
        # ----------------------------------------------------
 
        sheet["A3"] = (
            "Report Time"
        )
 
        sheet["B3"] = (
            datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S"
            )
        )
 
        # ----------------------------------------------------
        # Summary 데이터
        # ----------------------------------------------------
 
        summary = [
 
            (
                "Before PDF",
                before_file_count
            ),
 
            (
                "After PDF",
                after_file_count
            ),
 
            (
                "Matched PDF",
                matched_file_count
            ),
 
            (
                "Compared Pages",
                total_page_count
            ),
 
            (
                "Detected Changes",
                total_change_count
            ),
 
        ]
 
        start_row = 5
 
        for index, (
            label,
            value
        ) in enumerate(
            summary,
            start=start_row
        ):
 
            sheet.cell(
                row=index,
                column=1,
                value=label
            )
 
            sheet.cell(
                row=index,
                column=2,
                value=value
            )
 
            sheet.cell(
                row=index,
                column=1
            ).font = (
                self.header_font
            )
 
        # ----------------------------------------------------
        # 출력 파일
        # ----------------------------------------------------
 
        if output_path:
 
            sheet["A12"] = (
                "Output File"
            )
 
            sheet["B12"] = str(
                output_path
            )
 
        # ----------------------------------------------------
        # 열 너비
        # ----------------------------------------------------
 
        sheet.column_dimensions[
            "A"
        ].width = 25
 
        sheet.column_dimensions[
            "B"
        ].width = 35
 
 
    # ========================================================
    # CHANGE SHEET HEADER
    # ========================================================
 
    def create_change_sheet_header(
        self,
        workbook: Workbook
    ):
        """
        Changes Sheet의 Header를 생성한다.
        """
 
        sheet = workbook[
            "Changes"
        ]
 
        headers = [
 
            "ID",
 
            "Before File",
 
            "After File",
 
            "Before Page",
 
            "After Page",
 
            "Change Type",
 
            "Before Text",
 
            "After Text",
 
            "X",
 
            "Y",
 
            "Width",
 
            "Height",
 
            "Confidence",
 
            "Reason",
 
            "Before Image",
 
            "After Image",
 
        ]
 
        for column, header in enumerate(
            headers,
            start=1
        ):
 
            cell = sheet.cell(
                row=1,
                column=column,
                value=header
            )
 
            cell.font = (
                self.header_font
            )
 
            cell.alignment = (
                self.center_alignment
            )
 
            cell.border = (
                self.border
            )
 
        sheet.freeze_panes = (
            "A2"
        )
 
 
    # ========================================================
    # ADD CHANGE ROW
    # ========================================================
 
    def add_change_row(
        self,
        workbook: Workbook,
        change_id: int,
        before_file: str,
        after_file: str,
        before_page: int,
        after_page: int,
        region: ChangeRegion,
        text_change: Optional[
            TextChange
        ] = None,
        before_image_path: Optional[
            str
        ] = None,
        after_image_path: Optional[
            str
        ] = None,
    ):
        """
        하나의 변경점을 Excel에 추가한다.
        """
 
        sheet = workbook[
            "Changes"
        ]
 
        row = (
            sheet.max_row + 1
        )
 
        # ----------------------------------------------------
        # OCR 결과
        # ----------------------------------------------------
 
        before_text = ""
 
        after_text = ""
 
        change_type = (
            region.change_type
        )
 
        confidence = (
            region.confidence
        )
 
        reason = (
            region.reason
        )
 
        if text_change:
 
            before_text = (
                text_change.before_text
            )
 
            after_text = (
                text_change.after_text
            )
 
            change_type = (
                text_change.change_type
            )
 
            confidence = (
                text_change.confidence
            )
 
            reason = (
                text_change.reason
            )
 
        # ----------------------------------------------------
        # 기본 데이터
        # ----------------------------------------------------
 
        values = [
 
            change_id,
 
            before_file,
 
            after_file,
 
            before_page,
 
            after_page,
 
            change_type,
 
            before_text,
 
            after_text,
 
            region.x,
 
            region.y,
 
            region.width,
 
            region.height,
 
            round(
                confidence,
                4
            ),
 
            reason,
 
        ]
 
        for column, value in enumerate(
            values,
            start=1
        ):
 
            cell = sheet.cell(
                row=row,
                column=column,
                value=value
            )
 
            cell.alignment = (
                self.normal_alignment
            )
 
            cell.border = (
                self.border
            )
 
        # ----------------------------------------------------
        # 이미지 삽입
        # ----------------------------------------------------
 
        if before_image_path:
 
            self._insert_image(
                sheet,
                before_image_path,
                row,
                15
            )
 
        if after_image_path:
 
            self._insert_image(
                sheet,
                after_image_path,
                row,
                16
            )
 
        sheet.row_dimensions[
            row
        ].height = 100
 
 
    # ========================================================
    # INSERT IMAGE
    # ========================================================
 
    @staticmethod
    def _insert_image(
        sheet,
        image_path,
        row,
        column
    ):
        """
        Excel에 이미지를 삽입한다.
        """
 
        path = Path(
            image_path
        )
 
        if not path.exists():
 
            return
 
        try:
 
            image = ExcelImage(
                str(path)
            )
 
            image.width = 120
 
            image.height = 90
 
            cell = (
                f"{get_column_letter(column)}"
                f"{row}"
            )
 
            sheet.add_image(
                image,
                cell
            )
 
        except Exception:
 
            return
 
    # ========================================================
    # ADD PAGE RESULT
    # ========================================================
 
    def add_page_result(
        self,
        workbook: Workbook,
        before_file: str,
        after_file: str,
        before_page: int,
        after_page: int,
        result: ChangeDetectionResult,
        text_changes: Optional[
            List[TextChange]
        ] = None,
        crop_dir: Optional[str] = None,
        change_id_start: int = 1,
    ) -> int:
        """
        하나의 Before / After 페이지 비교 결과를
        Excel에 추가한다.
 
        반환값:
            다음에 사용할 change_id
        """
 
        if not result.success:
 
            return change_id_start
 
        if text_changes is None:
 
            text_changes = []
 
        # ----------------------------------------------------
        # OCR 결과를 region_id 기준으로 검색할 수 있도록
        # dictionary로 만든다.
        # ----------------------------------------------------
 
        text_change_map = {
            change.region_id: change
            for change in text_changes
        }
 
        change_id = (
            change_id_start
        )
 
        # ----------------------------------------------------
        # 변경 영역이 없는 페이지
        # ----------------------------------------------------
 
        if not result.regions:
 
            return change_id
 
        # ----------------------------------------------------
        # 각 변경 영역 처리
        # ----------------------------------------------------
 
        for region in result.regions:
 
            text_change = (
                text_change_map.get(
                    region.region_id
                )
            )
 
            before_image_path = None
 
            after_image_path = None
 
            # ------------------------------------------------
            # Crop 이미지 저장
            # ------------------------------------------------
 
            if crop_dir:
 
                before_image_path = (
                    self._save_region_image(
                        region.before_crop,
                        crop_dir,
                        change_id,
                        "before"
                    )
                )
 
                after_image_path = (
                    self._save_region_image(
                        region.after_crop,
                        crop_dir,
                        change_id,
                        "after"
                    )
                )
 
            # ------------------------------------------------
            # Excel row 추가
            # ------------------------------------------------
 
            self.add_change_row(
                workbook=workbook,
 
                change_id=change_id,
 
                before_file=before_file,
 
                after_file=after_file,
 
                before_page=before_page,
 
                after_page=after_page,
 
                region=region,
 
                text_change=text_change,
 
                before_image_path=(
                    before_image_path
                ),
 
                after_image_path=(
                    after_image_path
                ),
            )
 
            change_id += 1
 
        return change_id
 
 
    # ========================================================
    # SAVE REGION IMAGE
    # ========================================================
 
    @staticmethod
    def _save_region_image(
        image,
        output_dir,
        change_id,
        suffix
    ) -> Optional[str]:
        """
        변경 영역 Crop 이미지를 저장한다.
        """
 
        if image is None:
 
            return None
 
        if getattr(
            image,
            "size",
            0
        ) == 0:
 
            return None
 
        output_dir = Path(
            output_dir
        )
 
        output_dir.mkdir(
            parents=True,
            exist_ok=True
        )
 
        file_path = (
            output_dir
            /
            (
                f"change_"
                f"{change_id:04d}_"
                f"{suffix}.png"
            )
        )
 
        try:
 
            import cv2
 
            success = cv2.imwrite(
                str(file_path),
                image
            )
 
            if not success:
 
                return None
 
            return str(
                file_path
            )
 
        except Exception:
 
            return None
 
 
    # ========================================================
    # ADD ALIGNMENT RESULT
    # ========================================================
 
    def add_alignment_result(
        self,
        workbook: Workbook,
        before_file: str,
        after_file: str,
        before_page: int,
        after_page: int,
        scale_x: float,
        scale_y: float,
        translation_x: float,
        translation_y: float,
        alignment_score: float,
        alignment_status: str,
    ):
        """
        Auto Align 결과를 Excel의 Alignment Sheet에 추가한다.
        """
 
        sheet = workbook[
            "Alignment"
        ]
 
        # ----------------------------------------------------
        # Header가 없으면 생성
        # ----------------------------------------------------
 
        if sheet.max_row == 1:
 
            headers = [
 
                "Before File",
 
                "After File",
 
                "Before Page",
 
                "After Page",
 
                "Scale X",
 
                "Scale Y",
 
                "Translation X",
 
                "Translation Y",
 
                "Alignment Score",
 
                "Status",
 
            ]
 
            for column, header in enumerate(
                headers,
                start=1
            ):
 
                cell = sheet.cell(
                    row=1,
                    column=column,
                    value=header
                )
 
                cell.font = (
                    self.header_font
                )
 
                cell.alignment = (
                    self.center_alignment
                )
 
                cell.border = (
                    self.border
                )
 
        row = (
            sheet.max_row + 1
        )
 
        values = [
 
            before_file,
 
            after_file,
 
            before_page,
 
            after_page,
 
            round(
                scale_x,
                6
            ),
 
            round(
                scale_y,
                6
            ),
 
            round(
                translation_x,
                3
            ),
 
            round(
                translation_y,
                3
            ),
 
            round(
                alignment_score,
                4
            ),
 
            alignment_status,
 
        ]
 
        for column, value in enumerate(
            values,
            start=1
        ):
 
            cell = sheet.cell(
                row=row,
                column=column,
                value=value
            )
 
            cell.alignment = (
                self.center_alignment
            )
 
            cell.border = (
                self.border
            )
 
 
    # ========================================================
    # FORMAT SHEETS
    # ========================================================
 
    def format_sheets(
        self,
        workbook: Workbook
    ):
        """
        Excel 전체 Sheet의 열 너비와
        기본 표시 형식을 설정한다.
        """
 
        # ----------------------------------------------------
        # Summary
        # ----------------------------------------------------
 
        summary = workbook[
            "Summary"
        ]
 
        summary.column_dimensions[
            "A"
        ].width = 25
 
        summary.column_dimensions[
            "B"
        ].width = 35
 
        # ----------------------------------------------------
        # Changes
        # ----------------------------------------------------
 
        changes = workbook[
            "Changes"
        ]
 
        widths = {
 
            "A": 8,
 
            "B": 28,
 
            "C": 28,
 
            "D": 12,
 
            "E": 12,
 
            "F": 22,
 
            "G": 30,
 
            "H": 30,
 
            "I": 10,
 
            "J": 10,
 
            "K": 10,
 
            "L": 10,
 
            "M": 12,
 
            "N": 40,
 
            "O": 20,
 
            "P": 20,
 
        }
 
        for column, width in (
            widths.items()
        ):
 
            changes.column_dimensions[
                column
            ].width = width
 
        # ----------------------------------------------------
        # Alignment
        # ----------------------------------------------------
 
        alignment = workbook[
            "Alignment"
        ]
 
        alignment_widths = {
 
            "A": 28,
 
            "B": 28,
 
            "C": 12,
 
            "D": 12,
 
            "E": 12,
 
            "F": 12,
 
            "G": 18,
 
            "H": 18,
 
            "I": 18,
 
            "J": 20,
 
        }
 
        for column, width in (
            alignment_widths.items()
        ):
 
            alignment.column_dimensions[
                column
            ].width = width
 
 
    # ========================================================
    # SAVE WORKBOOK
    # ========================================================
 
    def save(
        self,
        workbook: Workbook,
        filename: Optional[str] = None,
    ) -> str:
        """
        Excel 파일을 저장한다.
        """
 
        if filename is None:
 
            timestamp = (
                datetime.now()
                .strftime(
                    "%Y%m%d_%H%M%S"
                )
            )
 
            filename = (
                f"DrawingCompare_"
                f"{timestamp}.xlsx"
            )
 
        filename = str(
            filename
        )
 
        if not filename.lower().endswith(
            ".xlsx"
        ):
 
            filename += ".xlsx"
 
        output_path = (
            self.output_dir
            /
            filename
        )
 
        # ----------------------------------------------------
        # Sheet Formatting
        # ----------------------------------------------------
 
        self.format_sheets(
            workbook
        )
 
        # ----------------------------------------------------
        # 저장
        # ----------------------------------------------------
 
        workbook.save(
            str(output_path)
        )
 
        return str(
            output_path
        )
 
    # ========================================================
    # GENERATE REPORT
    # ========================================================
 
    def generate_report(
        self,
        comparison_results: list,
        before_file_count: int = 0,
        after_file_count: int = 0,
        matched_file_count: int = 0,
        output_filename: Optional[str] = None,
    ) -> str:
        """
        전체 비교 결과를 Excel 보고서로 생성한다.
 
        comparison_results 형식:
 
        [
            {
                "before_file": "...",
                "after_file": "...",
                "before_page": 1,
                "after_page": 1,
                "result": ChangeDetectionResult,
                "text_changes": [TextChange, ...],
                "alignment": {
                    "scale_x": 1.0,
                    "scale_y": 1.0,
                    "translation_x": 0,
                    "translation_y": 0,
                    "score": 0.98,
                    "status": "OK"
                }
            },
            ...
        ]
        """
 
        # ----------------------------------------------------
        # Workbook 생성
        # ----------------------------------------------------
 
        workbook = (
            self.create_workbook()
        )
 
        # ----------------------------------------------------
        # Changes Header
        # ----------------------------------------------------
 
        self.create_change_sheet_header(
            workbook
        )
 
        # ----------------------------------------------------
        # 전체 페이지 수
        # ----------------------------------------------------
 
        total_page_count = len(
            comparison_results
        )
 
        # ----------------------------------------------------
        # 전체 변경점 수
        # ----------------------------------------------------
 
        total_change_count = 0
 
        for item in comparison_results:
 
            result = item.get(
                "result"
            )
 
            if result is None:
                continue
 
            if not result.success:
                continue
 
            total_change_count += len(
                result.regions
            )
 
        # ----------------------------------------------------
        # Summary
        # ----------------------------------------------------
 
        self.create_summary_sheet(
            workbook=workbook,
 
            before_file_count=(
                before_file_count
            ),
 
            after_file_count=(
                after_file_count
            ),
 
            matched_file_count=(
                matched_file_count
            ),
 
            total_page_count=(
                total_page_count
            ),
 
            total_change_count=(
                total_change_count
            ),
        )
 
        # ----------------------------------------------------
        # Crop 이미지 저장 폴더
        # ----------------------------------------------------
 
        crop_dir = (
            self.output_dir
            /
            "change_images"
        )
 
        crop_dir.mkdir(
            parents=True,
            exist_ok=True
        )
 
        # ----------------------------------------------------
        # 변경 ID
        # ----------------------------------------------------
 
        change_id = 1
 
        # ----------------------------------------------------
        # 모든 페이지 결과 처리
        # ----------------------------------------------------
 
        for item in comparison_results:
 
            before_file = str(
                item.get(
                    "before_file",
                    ""
                )
            )
 
            after_file = str(
                item.get(
                    "after_file",
                    ""
                )
            )
 
            before_page = int(
                item.get(
                    "before_page",
                    0
                )
            )
 
            after_page = int(
                item.get(
                    "after_page",
                    0
                )
            )
 
            result = item.get(
                "result"
            )
 
            text_changes = item.get(
                "text_changes",
                []
            )
 
            # ------------------------------------------------
            # ChangeDetectionResult
            # ------------------------------------------------
 
            if result is not None:
 
                change_id = (
                    self.add_page_result(
                        workbook=workbook,
 
                        before_file=(
                            before_file
                        ),
 
                        after_file=(
                            after_file
                        ),
 
                        before_page=(
                            before_page
                        ),
 
                        after_page=(
                            after_page
                        ),
 
                        result=result,
 
                        text_changes=(
                            text_changes
                        ),
 
                        crop_dir=(
                            str(crop_dir)
                        ),
 
                        change_id_start=(
                            change_id
                        ),
                    )
                )
 
            # ------------------------------------------------
            # Alignment 결과
            # ------------------------------------------------
 
            alignment = item.get(
                "alignment"
            )
 
            if alignment:
 
                self.add_alignment_result(
                    workbook=workbook,
 
                    before_file=(
                        before_file
                    ),
 
                    after_file=(
                        after_file
                    ),
 
                    before_page=(
                        before_page
                    ),
 
                    after_page=(
                        after_page
                    ),
 
                    scale_x=float(
                        alignment.get(
                            "scale_x",
                            1.0
                        )
                    ),
 
                    scale_y=float(
                        alignment.get(
                            "scale_y",
                            1.0
                        )
                    ),
 
                    translation_x=float(
                        alignment.get(
                            "translation_x",
                            0.0
                        )
                    ),
 
                    translation_y=float(
                        alignment.get(
                            "translation_y",
                            0.0
                        )
                    ),
 
                    alignment_score=float(
                        alignment.get(
                            "score",
                            0.0
                        )
                    ),
 
                    alignment_status=str(
                        alignment.get(
                            "status",
                            "UNKNOWN"
                        )
                    ),
                )
 
        # ----------------------------------------------------
        # 최종 저장
        # ----------------------------------------------------
 
        output_path = (
            self.save(
                workbook,
                output_filename
            )
        )
 
        return output_path
 
 
    # ========================================================
    # GENERATE SIMPLE REPORT
    # ========================================================
 
    def generate_simple_report(
        self,
        comparison_results: list,
        output_filename: Optional[str] = None,
    ) -> str:
        """
        파일 개수 등의 정보가 없는 경우에도
        간단하게 보고서를 생성할 수 있도록 한다.
        """
 
        # ----------------------------------------------------
        # 파일 목록 추출
        # ----------------------------------------------------
 
        before_files = set()
 
        after_files = set()
 
        for item in comparison_results:
 
            before_file = item.get(
                "before_file"
            )
 
            after_file = item.get(
                "after_file"
            )
 
            if before_file:
 
                before_files.add(
                    str(before_file)
                )
 
            if after_file:
 
                after_files.add(
                    str(after_file)
                )
 
        return self.generate_report(
 
            comparison_results=(
                comparison_results
            ),
 
            before_file_count=len(
                before_files
            ),
 
            after_file_count=len(
                after_files
            ),
 
            matched_file_count=len(
                set(
                    zip(
                        before_files,
                        after_files
                    )
                )
            ),
 
            output_filename=(
                output_filename
            ),
        )
 
 
# ============================================================
# DEFAULT REPORT GENERATOR
# ============================================================
 
_default_report_generator = (
    ReportGenerator()
)
 
 
def generate_excel_report(
    comparison_results: list,
    output_filename: Optional[str] = None,
) -> str:
    """
    외부 모듈에서 쉽게 사용할 수 있는
    Excel 생성 함수.
    """
 
    return (
        _default_report_generator
        .generate_simple_report(
            comparison_results,
            output_filename
        )
    )
 
# ============================================================
# VALIDATION
# ============================================================
 
def validate_report_generator() -> bool:
    """
    ReportGenerator가 정상적으로 동작하는지 확인한다.
    """
 
    try:
 
        generator = ReportGenerator()
 
        workbook = (
            generator.create_workbook()
        )
 
        if "Summary" not in workbook.sheetnames:
            return False
 
        if "Changes" not in workbook.sheetnames:
            return False
 
        if "Alignment" not in workbook.sheetnames:
            return False
 
        generator.create_change_sheet_header(
            workbook
        )
 
        generator.format_sheets(
            workbook
        )
 
        return True
 
    except Exception as exc:
 
        print(
            "ReportGenerator validation failed:"
        )
 
        print(exc)
 
        return False
 
 
# ============================================================
# TEST REPORT
# ============================================================
 
def create_test_report() -> Optional[str]:
    """
    실제 도면 데이터 없이
    빈 테스트 Excel을 생성한다.
 
    목적:
        openpyxl 설치 및
        ReportGenerator 코드가
        정상적으로 동작하는지 확인.
    """
 
    try:
 
        generator = (
            ReportGenerator(
                output_dir="output"
            )
        )
 
        comparison_results = []
 
        output_path = (
            generator.generate_report(
                comparison_results=(
                    comparison_results
                ),
 
                before_file_count=0,
 
                after_file_count=0,
 
                matched_file_count=0,
 
                output_filename=(
                    "DrawingCompare_test.xlsx"
                ),
            )
        )
 
        return output_path
 
    except Exception as exc:
 
        print(
            "Test report creation failed:"
        )
 
        print(exc)
 
        return None
 
 
# ============================================================
# MAIN TEST
# ============================================================
 
if __name__ == "__main__":
 
    print("=" * 60)
 
    print(
        "DrawingCompare H5"
    )
 
    print(
        "Report Generator Test"
    )
 
    print("=" * 60)
 
    # --------------------------------------------------------
    # Validation
    # --------------------------------------------------------
 
    valid = (
        validate_report_generator()
    )
 
    if valid:
 
        print(
            "✓ ReportGenerator 구조 정상"
        )
 
    else:
 
        print(
            "✗ ReportGenerator 구조 오류"
        )
 
    # --------------------------------------------------------
    # Test Excel
    # --------------------------------------------------------
 
    if valid:
 
        test_path = (
            create_test_report()
        )
 
        if test_path:
 
            print(
                "✓ Excel 테스트 파일 생성 성공"
            )
 
            print(
                f"파일 위치: {test_path}"
            )
 
        else:
 
            print(
                "✗ Excel 테스트 파일 생성 실패"
            )
 
    print("=" * 60)
 