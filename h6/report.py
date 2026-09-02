from pathlib import Path
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
import tempfile

def write_report(rows,path):
    path=Path(path); wb=Workbook(); ws=wb.active; ws.title='변경 결과'
    headers=['No','Before PDF','Before Page','After PDF','After Page','Type','Confidence','X','Y','Width','Height','Old Text','New Text','Before Image','After Image']
    ws.append(headers)
    tmp=Path(tempfile.mkdtemp(prefix='drawingcompare_h6_'))
    for i,(bp,bpg,ap,apg,c) in enumerate(rows,1):
        ws.append([i,bp,bpg,ap,apg,c.kind,round(c.confidence,3),c.x,c.y,c.w,c.h,c.old_text,c.new_text,'',''])
        for col,img in ((14,c.before_crop),(15,c.after_crop)):
            if img is None or img.size==0:continue
            p=tmp/f'{i}_{col}.png'; Image.fromarray(img[:,:,::-1]).save(p); x=XLImage(str(p)); x.width=min(320,x.width); x.height=min(220,x.height); ws.cell(i+1,col).value=''; ws.add_image(x,ws.cell(i+1,col).coordinate)
    for col,width in {'A':7,'B':28,'C':12,'D':28,'E':12,'F':14,'G':12,'H':10,'I':10,'J':10,'K':10,'L':28,'M':28,'N':32,'O':32}.items():ws.column_dimensions[col].width=width
    ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions; wb.save(path); return path
