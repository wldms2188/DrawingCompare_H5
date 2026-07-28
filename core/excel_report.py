import os
 
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
 
 
class ExcelReport:
 
    def create(self, changes, output_path):
 
        wb = Workbook()
        ws = wb.active
        ws.title = "Drawing Compare"
 
        headers = [
            "No",
            "Page",
            "Type",
            "Before",
            "After",
            "Before Image",
            "After Image"
        ]
 
        for col, title in enumerate(headers, start=1):
 
            cell = ws.cell(row=1, column=col)
            cell.value = title
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
 
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 30
        ws.column_dimensions["E"].width = 30
        ws.column_dimensions["F"].width = 25
        ws.column_dimensions["G"].width = 25
 
        for row, change in enumerate(changes, start=2):
 
            ws.row_dimensions[row].height = 90
 
            ws.cell(row=row, column=1).value = change.id
            ws.cell(row=row, column=2).value = change.page
            ws.cell(row=row, column=3).value = change.change_type
            ws.cell(row=row, column=4).value = change.before_text
            ws.cell(row=row, column=5).value = change.after_text
 
            if os.path.exists(change.before_image_path):
 
                img = Image(change.before_image_path)
                img.width = 150
                img.height = 80
 
                ws.add_image(img, f"F{row}")
 
            if os.path.exists(change.after_image_path):
 
                img = Image(change.after_image_path)
                img.width = 150
                img.height = 80
 
                ws.add_image(img, f"G{row}")
 
        os.makedirs(
            os.path.dirname(output_path),
            exist_ok=True
        )
 
        wb.save(output_path)
 
        print(f"Excel 저장 완료 : {output_path}")